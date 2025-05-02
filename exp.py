# Import libraries
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import streamlit as st
from openpyxl import load_workbook

# Set page config
st.set_page_config(layout="wide")
st.title("COVID-19 Data Analysis Dashboard")

# Display header
st.header("Dataset Exploration")
st.write("This dashboard analyzes COVID-19 testing data across different regions.")

# Read dataset
@st.cache_data  # Cache for performance
def load_data():
    covid_excel = pd.ExcelFile(r"C:\Users\user\Downloads\archive\covid_testing.xlsx")
    return {sheet: covid_excel.parse(sheet) for sheet in covid_excel.sheet_names}

all_sheets = load_data()

# Display sheet info
st.subheader("Original Sheets in Dataset")
for name, df in all_sheets.items():
    st.write(f"{name}: {df.shape[0]} rows, {df.shape[1]} columns")

# Rename sheets
st.subheader("Renaming Sheets")
new_names = {
    "TimeSeries_KeyIndicators": "KeyMetrics",
    "TimeSeries_KeyIndicators_Detail": "DetailedMetrics",
    "TimeSeries_Action_Screen": "Actions",
    "TimeSeries_Action_Call": "Calls",
    "TimeSeries_COVID_News": "News",
    "TimeSeries_NIH_Response": "NIH_Response",
    "TimeSeries_NIH_Risk": "RiskLevels",
    "TimeSeries_Province_Response": "ProvinceData",
    "TimeSeries_Helpline_Calls": "Helpline",
    "TimeSeries_Quarantine_Details": "Quarantine",
}

# Save renamed version
@st.cache_data
def rename_sheets():
    wb = load_workbook(r"C:\Users\user\Downloads\archive\covid_testing.xlsx")
    for old_name in new_names:
        if old_name in wb.sheetnames:
            wb[old_name].title = new_names[old_name]
    wb.save("covid_testing_renamed.xlsx")
    return pd.read_excel("covid_testing_renamed.xlsx", sheet_name=None)

covid_excel = rename_sheets()

# Display renamed sheets
st.write("Sheets after renaming:")
st.write(list(covid_excel.keys()))

# Load main dataframe
df1 = covid_excel["KeyMetrics"].copy()
df1.drop(columns="Unnamed: 0", inplace=True)

# Rename columns
df1.rename(columns={
    "Home Quarantine": "quarantine",
    "Cumulative  Test positive": "cum_test_positive",
    "Cumulative  tests performed": "cum_test_performed",
    "New  (last 24 hrs)": "new_test_24hrs",
    "Still admitted": "still_admitted",
    "Tests  performed in last 24 hours": "test_24hrs",
}, inplace=True)

# Data cleaning
st.subheader("Data Cleaning")
df1.fillna({"quarantine": df1["quarantine"].mean()}, inplace=True)
df1["Region"] = df1["Region"].replace({"KPTD": "KP"})

# Show cleaned data
st.write("Cleaned Data Preview:")
st.dataframe(df1.head())

# Outlier handling
st.subheader("Outlier Analysis")
st.write("Original Statistics:")
st.write(df1["Cumulative"].describe())

# Visualization tabs
tab1, tab2 = st.tabs(["Distribution", "Boxplot"])

with tab1:
    fig1 = plt.figure(figsize=(10, 6))
    sns.histplot(data=df1, x="Cumulative", kde=True)
    st.pyplot(fig1)

with tab2:
    fig2 = plt.figure(figsize=(10, 6))
    sns.boxplot(data=df1, y="Cumulative")
    st.pyplot(fig2)

# Outlier removal
st.subheader("After Outlier Removal")
df1 = df1[(df1["Cumulative"] > 0) & (df1["Cumulative"] < 14000)]

Q1 = df1["Cumulative"].quantile(0.25)
Q3 = df1["Cumulative"].quantile(0.75)
IQR = Q3 - Q1
df1 = df1 [~((df1["Cumulative"] < (Q1 - 1.5 * IQR)) | (df1["Cumulative"] > (Q3 + 1.5 * IQR)))]

st.write("Updated Statistics:")
st.write(df1["Cumulative"].describe())

# Final visualizations
col1, col2 = st.columns(2)
with col1:
    fig3 = plt.figure()
    sns.histplot(data=df1, x="Cumulative", kde=True)
    st.pyplot(fig3)

with col2:
    fig4 = plt.figure()
    sns.boxplot(data=df1, y="Cumulative")
    st.pyplot(fig4)

# Add download button
st.download_button(
    label="Download Cleaned Data",
    data=df1.to_csv().encode('utf-8'),
    file_name='cleaned_covid_data.csv',
    mime='text/csv'
)

st.success("Analysis complete!")
